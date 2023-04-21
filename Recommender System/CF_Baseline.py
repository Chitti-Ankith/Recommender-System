#This code calculates RMSE,Spearman Correlation coefficient and Precision at Top K
import timeit
import random
import numpy as np
import multiprocessing
import xlrd
import math
from scipy import sparse 
from openpyxl import load_workbook
from sklearn.model_selection import train_test_split


npr=4

def mag(x): 
    return math.sqrt(sum(i**2 for i in x))

def adder(x,y):
	s1=(x*(x+1))/2
	s2=(y*(y+1))/2
	return s2-s1

def find_averages(original_sparse_matrix,bx,by,max_row,max_col):
	total_sum=0
	total_entries=0
	for i in range(max_row):
		sum_of_elements_in_row=0
		row_start_index_in_data_list=original_sparse_matrix.indptr[i]
		row_end_index_in_data_list=original_sparse_matrix.indptr[i+1]
		no_of_elements_in_row=row_end_index_in_data_list-row_start_index_in_data_list
		row_data=original_sparse_matrix.data[row_start_index_in_data_list:row_end_index_in_data_list]

		for d in range(len(row_data)):
			sum_of_elements_in_row+=row_data[d]
		bx.append(sum_of_elements_in_row/no_of_elements_in_row)

		total_sum+=sum_of_elements_in_row
		total_entries+=len(row_data)

	u=total_sum/total_entries

	original_sparse_matrix_transpose=original_sparse_matrix.transpose()
	for i in range(max_col):
		sum_of_elements_in_row=0
		row_start_index_in_data_list=original_sparse_matrix_transpose.indptr[i]
		row_end_index_in_data_list=original_sparse_matrix_transpose.indptr[i+1]
		no_of_elements_in_row=row_end_index_in_data_list-row_start_index_in_data_list
		row_data=original_sparse_matrix_transpose.data[row_start_index_in_data_list:row_end_index_in_data_list]

		for d in range(len(row_data)):
			sum_of_elements_in_row+=row_data[d]
		by.append(sum_of_elements_in_row/no_of_elements_in_row)

	return u

def average_out_ratings_and_row_normalize(maxm,sparse_matrix):
	for i in range(maxm):
		sum_of_elements_in_row=0
		row_start_index_in_data_list=sparse_matrix.indptr[i]
		row_end_index_in_data_list=sparse_matrix.indptr[i+1]
		no_of_elements_in_row=row_end_index_in_data_list-row_start_index_in_data_list
		row_data=sparse_matrix.data[row_start_index_in_data_list:row_end_index_in_data_list]

		for d in range(len(row_data)):
			sum_of_elements_in_row+=row_data[d]

		average=sum_of_elements_in_row/no_of_elements_in_row

		for d in range(len(row_data)):
			row_data[d]=row_data[d]-average

		row_data=sparse_matrix.data[row_start_index_in_data_list:row_end_index_in_data_list]
		row_length=mag(row_data)

		for d in range(len(row_data)):
			row_data[d]/=row_length

def spearman_rank_correlation_calculator(x,y):
	s1=sum(i for i in x)
	s2=sum(i for i in y)
	s11=sum(i**2 for i in x)
	s22=sum(i**2 for i in y)
	s12=0
	for i in range(len(x)):
		s12+=x[i]*y[i]
		
	num=(len(x)*s12-s1*s2)
	den=((len(x)*s11 - s1*s1)*(len(x)*s22 - s2*s2))**0.5
	cor=num/den
	print(cor)


def average_precision_at_topK(t2,t3,t4,max_row):
	random_users_list=random.sample(range(max_row),1000)
	precattopk=0
	den=0
	for i in range(len(random_users_list)):
		try:
			firstocc=t2.index(random_users_list[i])
			lastocc=len(t2)-t2[::-1].index(random_users_list[i])-1
			if((lastocc-firstocc)>=10):
				recommend=0
				relevant=0
				for j in range(firstocc,firstocc+10):
					if(t3[j]>3.5):
						recommend+=1
						if(t4[j]>3.5):
							relevant+=1
				if(recommend!=0):
					precattopk+=(relevant/recommend)
					den+=1
		except ValueError:
			continue
	average_precision_at_top_K=(precattopk/den)
	print(average_precision_at_top_K)



#function for predicting rating of a particular item for a user using user-user collaborative filtering
def rating_calculator(unq,ro,co,rat,spars_matrix,spars_matrix_transpose,max_row_index,original_matrix,lock,us,ac,pre,usrin,actin,precin,u,bx,by):

	xx=int(unq*(len(ro)/npr))
	yy=int((unq+1)*(len(ro)/npr))
	
	#converting to dense matrix
	spars_matrix=spars_matrix.toarray()
	spars_matrix_transpose=spars_matrix_transpose.toarray()
	#finding simlarity scores
	sim_score_matrix_dense=np.dot(spars_matrix,spars_matrix_transpose)
	original_matrix_dense=original_matrix.toarray()
	diff=[]


	for j in range(xx,yy):
		row_index=int(ro[j])
		col_index=int(co[j])

		weighted_average=-1
		prod=0
		sum_of_weights=0
		sim_score_list={}

		dx=0

		for i in range(max_row_index):
			ith_row_sim_score=0
			if(i!=row_index):
				if(original_matrix_dense[i][col_index]>=1):
					if(sim_score_matrix_dense[row_index][i]>=0):
						ith_row_sim_score=sim_score_matrix_dense[row_index][i]
						dx+=1
				sim_score_list[i]=ith_row_sim_score	
			
		sorted_by_value=sorted(sim_score_list.items(), key=lambda kv: kv[1], reverse=True)

		for y in range(min(10,dx)):

			rowy=sorted_by_value[y][0]
			weight=sorted_by_value[y][1]
			nthg=original_matrix_dense[rowy][col_index]
			nthg-=(bx[rowy]+by[col_index]-u)
			sum_of_weights+=weight
			prod+=(nthg*weight)

		if(sum_of_weights!=0):
			weighted_average=prod/sum_of_weights
			weighted_average+=(bx[row_index]+by[col_index]-u)
			diff.append(rat[j]-weighted_average)

		#writing to shared memory
		lock.acquire()

		us[usrin.value]=row_index
		usrin.value+=1
		ac[actin.value]=rat[j]
		actin.value+=1
		pre[precin.value]=weighted_average
		precin.value+=1

		lock.release()
		#calculating RMSE
	RMSE=math.sqrt(sum(i**2 for i in diff)/len(diff))
	
	#writing RMSE calculated by process back to excel file

	lock.acquire()
	wb=load_workbook("answer.xlsx")
	ws=wb.active
	
	cell1="A"+str(unq+1)
	cell2="B"+str(unq+1)	

	ws[cell1]=RMSE
	ws[cell2]=len(diff)
	wb.save("answer.xlsx")

	lock.release()


def feed_training_and_testing_data_set(pos_train,pos_test,rating_train,rating_test,max_row,max_col):

	row_train,col_train=zip(*pos_train)
	row_test,col_test=zip(*pos_test)

	#declaring variables and arrays in shared memory to store user,actual rating and predicted rating
	usr=multiprocessing.Array('i',len(row_test))
	act=multiprocessing.Array('d',len(row_test))
	prec=multiprocessing.Array('d',len(row_test))

	usrindx=multiprocessing.Value('i')
	actindx=multiprocessing.Value('i')
	precindx=multiprocessing.Value('i')

	usrindx.value=0
	actindx.value=0
	precindx.value=0

	row_train=np.asarray(row_train)
	col_train=np.asarray(col_train)
	rating_train=np.asarray(rating_train)

	#creating sparse matrix from training dataset
	original_sparse_matrix=sparse.csr_matrix((rating_train,(row_train,col_train)),shape=(max_row,max_col))
	bx=[]
	by=[]
	#finding overall average rating,users and items average ratings for implementing global and local effects in model
	u=find_averages(original_sparse_matrix,bx,by,max_row,max_col)

	sparse_matrix=original_sparse_matrix.copy()
	#calling function to average out ratings of a user and then row normalize
	average_out_ratings_and_row_normalize(max_row,sparse_matrix)
	sparse_matrix_transpose=sparse_matrix.transpose()

	lock=multiprocessing.Lock()#lock variable from multiprocessing module
	
	start=timeit.default_timer()

	pr={}
	for k in range(npr):
		#generating multiple processes 
		pr[k]=multiprocessing.Process(target=rating_calculator, args=(k,row_test,col_test,rating_test,sparse_matrix,sparse_matrix_transpose,max_row,original_sparse_matrix,lock,usr,act,prec,usrindx,actindx,precindx,u,bx,by)) 
		pr[k].start()

	for k in range(npr):
		pr[k].join()

	stop=timeit.default_timer()
	time_taken=(stop-start)/len(row_test)
	print('average time taken for a prediction of rating of an item for an user is ',time_taken)

	#finding RMSE from RMSE calculated by various processes after reading what these processes have written to excel file

	loc=("answer.xlsx")
	wb=xlrd.open_workbook(loc) 	
	sheet=wb.sheet_by_index(0)
	tn=0
	p=0
	for i in range(sheet.nrows): 
		r=sheet.cell_value(i,0)
		n=sheet.cell_value(i,1)
		tn+=n
		p+=(r**2)*n
	RMSE_TOTAL=math.sqrt(p/tn)
	print(RMSE_TOTAL)

	
	#average precision at topK
	t1=list(zip(usr,prec,act))
	t1.sort(key=lambda x:(x[0],-x[1]))
	t2,t3,t4=zip(*t1)
	average_precision_at_topK(t2,t3,t4,max_row)

	#spearman rank correlation
	t5=list(zip(t3,t4))
	t5.sort(key=lambda x:(x[0]))
	t3,t4=zip(*t5)
	
	ranked_t3=[]
	nn=1
	for i in range(len(t3)):
		ranked_t3.append(nn)
		nn+=1

	ranked_t4=[]
	numb=[0,0,0,0,0]
	for i in range(len(t4)):
		numb[int(t4[i])-1]+=1
	
	r1=adder(0,numb[0])/numb[0]
	r2=adder(numb[0],numb[1]+numb[0])/numb[1]	
	r3=adder(numb[0]+numb[1],numb[2]+numb[1]+numb[0])/numb[2]
	r4=adder(numb[0]+numb[1]+numb[2],numb[3]+numb[2]+numb[1]+numb[0])/numb[3]
	r5=adder(numb[0]+numb[1]+numb[2]+numb[3],numb[4]+numb[3]+numb[2]+numb[1]+numb[0])/numb[4]
	for i in range(len(t4)):
		if(t4[i]==1):
			ranked_t4.append(r1)
		elif(t4[i]==2):
			ranked_t4.append(r2)
		elif(t4[i]==3):
			ranked_t4.append(r3)
		elif(t4[i]==4):
			ranked_t4.append(r4)
		elif(t4[i]==5):
			ranked_t4.append(r5)

	spearman_rank_correlation_calculator(ranked_t3,ranked_t4)

	
def main():

	loc=("ratings.xlsx") 
 	
	row_list=[]
	col_list=[]
	rating_list=[]

	max_row_index=0
	max_col_index=0

	wb=xlrd.open_workbook(loc) 	
	sheet=wb.sheet_by_index(0) 

	#reading data from excel file
	for i in range(sheet.nrows): 
		t1=sheet.cell_value(i,0)
		t2=sheet.cell_value(i,1)
		row_list.append(t1-1)
		col_list.append(t2-1)
		rating_list.append(sheet.cell_value(i,2))
		if t1>max_row_index:
			max_row_index=t1
		if t2>max_col_index:
			max_col_index=t2

	max_row_index=int(max_row_index)
	max_col_index=int(max_col_index)

	pos_list=list(zip(row_list,col_list))
	
	#splitting dataset into training and testing dataset
	pos_train,pos_test,rating_train,rating_test=train_test_split(pos_list,rating_list,test_size=0.20,random_state=11)

	#calling a function to predict ratings and calculate model evaluation metrics
	feed_training_and_testing_data_set(pos_train,pos_test,rating_train,rating_test,max_row_index,max_col_index)



if __name__ == '__main__':
		if(main()==-1):
			sys.exit(0)